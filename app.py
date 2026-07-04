"""
GST Reconciliation System – Single Page Edition
AI-powered invoice matching between GSTR-2B and Purchase Register
Prepared & Developed by Karthik LVN | 9849270702
"""

import io
import datetime
import pandas as pd
import streamlit as st

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GST Reconciliation – Karthik LVN",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={
        "About": "GST Reconciliation v2.0\nPrepared & Developed by Karthik LVN | 9849270702",
    },
)

# ── Bootstrap ──────────────────────────────────────────────────────────────
from modules.utils import ensure_directories, setup_logging
from modules.audit import initialize_audit_db, log_event

ensure_directories()
initialize_audit_db()
logger = setup_logging()

# ── Global CSS ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Hide sidebar completely */
[data-testid="stSidebar"]          { display: none !important; }
[data-testid="collapsedControl"]   { display: none !important; }

/* Light page background */
.stApp {
    background: #F1F5F9;
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif !important;
}

/* Centre & max-width main content */
.main .block-container {
    max-width: 960px !important;
    padding: 24px 24px 40px 24px !important;
    margin: 0 auto !important;
}

/* Hide Streamlit chrome */
#MainMenu { visibility: hidden; }
footer    { visibility: hidden; }
header    { visibility: hidden; }
.stDeployButton { display: none; }

/* ── Top-LEFT contact badge ── */
#karthik-badge {
    position: fixed;
    top: 8px;
    left: 16px;
    z-index: 9999;
    background: rgba(15,23,42,0.95);
    border: 1px solid rgba(0,212,255,0.4);
    border-radius: 24px;
    padding: 4px 16px;
    display: flex;
    align-items: center;
    gap: 10px;
    font-family: 'Segoe UI', sans-serif;
    backdrop-filter: blur(10px);
    box-shadow: 0 2px 12px rgba(0,0,0,0.25);
}
.kb-name  { font-size: 0.76rem; font-weight: 700; color: #00D4FF; }
.kb-sep   { color: rgba(255,255,255,0.2); }
.kb-phone { font-size: 0.72rem; color: #A78BFA; font-weight: 600; }

/* ── Upload drop-zone look ── */
[data-testid="stFileUploadDropzone"] {
    border: 2px dashed #CBD5E1 !important;
    border-radius: 10px !important;
    background: #FFFFFF !important;
    min-height: 56px !important;
}
[data-testid="stFileUploadDropzone"]:hover {
    border-color: #6366F1 !important;
    background: #F5F3FF !important;
}

/* ── Reconcile button (big, purple-blue) ── */
div[data-testid="stButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, #4F46E5, #7C3AED) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 10px !important;
    font-size: 1rem !important;
    font-weight: 700 !important;
    padding: 12px 24px !important;
    letter-spacing: 0.3px !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 4px 16px rgba(79,70,229,0.35) !important;
}
div[data-testid="stButton"] > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #4338CA, #6D28D9) !important;
    box-shadow: 0 6px 20px rgba(79,70,229,0.5) !important;
    transform: translateY(-1px) !important;
}

/* Download button (green) */
.dl-btn > button {
    background: linear-gradient(135deg, #059669, #10B981) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 0.95rem !important;
}

/* ── Metrics ── */
[data-testid="metric-container"] {
    background: #FFFFFF !important;
    border: 1px solid #E2E8F0 !important;
    border-radius: 12px !important;
    padding: 14px 16px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
}
[data-testid="stMetricValue"] { color: #0F172A !important; font-weight: 800 !important; }
[data-testid="stMetricLabel"] { color: #64748B !important; font-size: 0.78rem !important; }

/* ── Download button styling ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #059669, #10B981) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 0.95rem !important;
    padding: 10px 20px !important;
    box-shadow: 0 4px 12px rgba(5,150,105,0.3) !important;
    transition: all 0.2s ease !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #047857, #059669) !important;
    box-shadow: 0 6px 16px rgba(5,150,105,0.45) !important;
    transform: translateY(-1px) !important;
}

/* Success/error alerts */
.stAlert { border-radius: 10px !important; }

/* Section labels */
.sec-label {
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 1.5px;
    color: #94A3B8;
    margin-bottom: 10px;
    text-transform: uppercase;
}

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #F1F5F9; }
::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #94A3B8; }
</style>
""", unsafe_allow_html=True)

# ── Top-right name badge ───────────────────────────────────────────────────
st.markdown("""
<div id="karthik-badge">
    <span class="kb-name">&#128100; Karthik LVN</span>
    <span class="kb-sep">|</span>
    <span class="kb-phone">&#128222;&nbsp;9849270702</span>
</div>
""", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────
_defaults = {
    "recon_results": None,
    "master_df":     None,
    "gstr_name":     None,
    "pr_name":       None,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Header banner ──────────────────────────────────────────────────────────
st.markdown("""
<div style="background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%);
     border-radius: 16px; padding: 28px 36px; margin-bottom: 24px;
     border: 1px solid rgba(255,255,255,0.06);
     box-shadow: 0 8px 32px rgba(0,0,0,0.18);">
    <div style="display:flex; align-items:center; gap:14px; margin-bottom:8px;">
        <span style="font-size:2rem;">⚡</span>
        <span style="font-size:1.7rem; font-weight:800; color:#FFFFFF;
               letter-spacing:-0.5px;">GST Reconciliation</span>
        <span style="background:rgba(99,102,241,0.25); color:#A5B4FC; font-size:0.68rem;
               font-weight:700; padding:3px 10px; border-radius:20px;
               border:1px solid rgba(99,102,241,0.4); letter-spacing:0.5px;">v2.0</span>
    </div>
    <div style="color:#94A3B8; font-size:0.9rem; padding-left:4px; margin-bottom:4px;">
        AI-powered invoice matching between GSTR-2B and Purchase Register
    </div>
    <div style="color:#475569; font-size:0.78rem; padding-left:4px;">
        Prepared &amp; Developed by
        <strong style="color:#00D4FF;">Karthik LVN</strong>
        &nbsp;&middot;&nbsp; &#128222; 9849270702
    </div>
</div>
""", unsafe_allow_html=True)


# ===========================================================================
# Helpers
# ===========================================================================

def read_uploaded_file(file) -> pd.DataFrame:
    """Read Excel or CSV upload into a DataFrame."""
    try:
        name = file.name.lower()
        if name.endswith(".csv"):
            return pd.read_csv(file)
        else:
            return pd.read_excel(file, engine="openpyxl")
    except Exception as e:
        st.error(f"Could not read **{file.name}**: {e}")
        return pd.DataFrame()


def auto_process(df: pd.DataFrame, source_tag: str) -> pd.DataFrame:
    """Silently auto-detect columns, apply mapping & clean – no user prompts."""
    try:
        from modules.mapping import auto_detect_columns, apply_mapping, clean_dataframe
        cols = list(df.columns)
        mapping = auto_detect_columns(cols) or {}
        # Fallback: if mapping is empty, try direct match to standard names
        if not mapping:
            STD = ["vendor_name","gstin","invoice_number","invoice_date",
                   "taxable_value","cgst","sgst","igst","cess","total_gst","invoice_value"]
            mapping = {c: c for c in STD if c in cols}
        mapped  = apply_mapping(df, mapping, source_tag=source_tag)
        cleaned, _ = clean_dataframe(mapped, mapping)
        return cleaned
    except Exception as exc:
        logger.warning(f"Auto-process failed for {source_tag}: {exc}")
        return df


def build_master(recon_results: dict) -> pd.DataFrame:
    """Flatten all result sets into one table with a Status column."""
    WANT = ["vendor_name", "gstin", "invoice_number", "invoice_date",
            "taxable_value", "cgst", "sgst", "igst", "total_gst", "invoice_value"]

    def _extract(df, status, pref):
        if df is None or (hasattr(df, "empty") and df.empty):
            return pd.DataFrame()
        rec = {}
        for c in WANT:
            pk = c + pref
            if pref and pk in df.columns:
                rec[c] = df[pk].values
            elif c in df.columns:
                rec[c] = df[c].values
        if not rec:
            return pd.DataFrame()
        out = pd.DataFrame(rec)
        out.insert(0, "Status", status)
        for extra in ["Match Reason", "Similarity %", "confidence"]:
            if extra in df.columns:
                out[extra] = df[extra].values
        return out

    frames = [
        _extract(recon_results.get("matched"),           "Fully Matched",     "_gstr2b"),
        _extract(recon_results.get("missing_in_books"),  "Missing in Books",   ""),
        _extract(recon_results.get("fuzzy_candidates"),  "Near Match",         "_gstr2b"),
        _extract(recon_results.get("missing_in_gstr2b"), "Missing in GSTR-2B", ""),
    ]
    valid = [f for f in frames if isinstance(f, pd.DataFrame) and not f.empty]
    return pd.concat(valid, ignore_index=True) if valid else pd.DataFrame()


def make_template_bytes() -> bytes:
    """Build a two-sheet Excel template (GSTR-2B + Purchase Register)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    COLS = ["vendor_name","gstin","invoice_number","invoice_date",
            "taxable_value","cgst","sgst","igst","cess","total_gst","invoice_value"]
    SAMPLES = {
        "vendor_name":   ["ABC Traders Pvt Ltd","XYZ Industries Pvt Ltd","PQR Enterprises"],
        "gstin":         ["29ABCDE1234F1Z5","27XYZPQ5678G1Z3","33PQRST9012H1Z7"],
        "invoice_number":["INV/2024/001","INV/2024/002","INV/2024/003"],
        "invoice_date":  ["01-04-2024","05-04-2024","10-04-2024"],
        "taxable_value": [100000,250000,75000],
        "cgst":          [9000,22500,6750],
        "sgst":          [9000,22500,6750],
        "igst":          [0,0,0],
        "cess":          [0,0,0],
        "total_gst":     [18000,45000,13500],
        "invoice_value": [118000,295000,88500],
    }
    wb = openpyxl.Workbook()
    for sheet in ["GSTR-2B","Purchase Register"]:
        ws = wb.create_sheet(sheet)
        hf  = PatternFill("solid", fgColor="1E3A5F")
        hft = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        ctr = Alignment(horizontal="center")
        for ci, col in enumerate(COLS, 1):
            c = ws.cell(1, ci, col); c.font = hft; c.fill = hf; c.alignment = ctr
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20
        for ri in range(3):
            for ci, col in enumerate(COLS, 1):
                vals = SAMPLES.get(col, ["","",""])
                ws.cell(ri + 2, ci, vals[ri] if ri < len(vals) else "")
    del wb["Sheet"]
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ===========================================================================
# SECTION 1 — Template
# ===========================================================================
st.markdown('<div class="sec-label">Template</div>', unsafe_allow_html=True)

t_col, _ = st.columns([2, 8])
with t_col:
    st.download_button(
        "⬇ Download Template",
        data=make_template_bytes(),
        file_name="GST_Recon_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_tmpl",
        use_container_width=True,
    )

st.markdown("<br>", unsafe_allow_html=True)

# ===========================================================================
# SECTION 2 — Upload Files
# ===========================================================================
st.markdown('<div class="sec-label">Upload Files</div>', unsafe_allow_html=True)

ul1, ul2 = st.columns(2, gap="large")

with ul1:
    st.markdown("""
    <div style="background:#FFFFFF; border:2px dashed #CBD5E1; border-radius:14px;
         padding:22px 16px; text-align:center; margin-bottom:8px;">
        <div style="font-size:2.2rem; margin-bottom:6px;">📋</div>
        <div style="font-weight:700; color:#1E293B; font-size:0.95rem;">GSTR-2B File</div>
        <div style="color:#94A3B8; font-size:0.76rem; margin-top:4px;">
            Upload your GSTR-2B Excel file (.xlsx)
        </div>
    </div>""", unsafe_allow_html=True)
    gstr_file = st.file_uploader(
        "GSTR-2B",
        type=["xlsx","xls","csv"],
        label_visibility="collapsed",
        key="gstr_upload",
    )

with ul2:
    st.markdown("""
    <div style="background:#FFFFFF; border:2px dashed #CBD5E1; border-radius:14px;
         padding:22px 16px; text-align:center; margin-bottom:8px;">
        <div style="font-size:2.2rem; margin-bottom:6px;">📄</div>
        <div style="font-weight:700; color:#1E293B; font-size:0.95rem;">Purchase Register</div>
        <div style="color:#94A3B8; font-size:0.76rem; margin-top:4px;">
            Upload your Purchase Register Excel file (.xlsx)
        </div>
    </div>""", unsafe_allow_html=True)
    pr_file = st.file_uploader(
        "Purchase Register",
        type=["xlsx","xls","csv"],
        label_visibility="collapsed",
        key="pr_upload",
    )

# Status banners
if gstr_file or pr_file:
    b1, b2 = st.columns(2)
    with b1:
        if gstr_file:
            st.success(f"✓ GSTR-2B uploaded: **{gstr_file.name}**")
        else:
            st.warning("Waiting for GSTR-2B file…")
    with b2:
        if pr_file:
            st.success(f"✓ Purchase Register uploaded: **{pr_file.name}**")
        else:
            st.warning("Waiting for Purchase Register file…")

st.markdown("<br>", unsafe_allow_html=True)

# ===========================================================================
# SECTION 3 — Reconcile Button
# ===========================================================================
_, rbtn, _ = st.columns([2, 6, 2])
with rbtn:
    do_reconcile = st.button(
        "⚡  Reconcile Data with AI",
        type="primary",
        use_container_width=True,
        key="reconcile_btn",
        disabled=(gstr_file is None or pr_file is None),
    )

if do_reconcile and gstr_file and pr_file:
    progress = st.progress(0, "Reading files…")
    try:
        gstr_raw = read_uploaded_file(gstr_file)
        pr_raw   = read_uploaded_file(pr_file)

        if gstr_raw.empty or pr_raw.empty:
            st.error("One or both files appear empty or could not be parsed.")
            progress.empty()
        else:
            progress.progress(20, "Auto-detecting columns…")
            gstr_proc = auto_process(gstr_raw.copy(), "GSTR2B")
            pr_proc   = auto_process(pr_raw.copy(),   "PR")

            progress.progress(50, "Running AI reconciliation…")
            from modules.matching import run_reconciliation
            results = run_reconciliation(pr_proc, gstr_proc)

            progress.progress(85, "Building results table…")
            master = build_master(results)

            st.session_state["recon_results"] = results
            st.session_state["master_df"]     = master
            st.session_state["gstr_name"]     = gstr_file.name
            st.session_state["pr_name"]       = pr_file.name
            log_event("RECONCILE", f"Complete: {results.get('stats', {})}")

            progress.progress(100, "Done!")
            import time; time.sleep(0.4)
            progress.empty()
            st.rerun()

    except Exception as exc:
        st.error(f"Reconciliation failed: {exc}")
        logger.exception("Reconciliation error")
        progress.empty()

# ===========================================================================
# SECTION 4 — Results (shown after reconciliation)
# ===========================================================================
recon_results = st.session_state.get("recon_results")
master_df     = st.session_state.get("master_df")

if recon_results is not None and master_df is not None and not master_df.empty:
    stats = recon_results.get("stats", {})

    st.divider()
    st.markdown('<div class="sec-label">Summary</div>', unsafe_allow_html=True)

    # ── KPI row ─────────────────────────────────────────────────────────────
    mc1, mc2, mc3, mc4, mc5, mc6 = st.columns(6)
    mc1.metric("Total GSTR-2B",    f"{stats.get('total_gstr2b', 0):,}")
    mc2.metric("Total PR",         f"{stats.get('total_pr', 0):,}")
    mc3.metric("Fully Matched",    f"{stats.get('total_matched', 0):,}")
    mc4.metric("Near Match",       f"{stats.get('fuzzy_candidates_count', 0):,}")
    mc5.metric("Missing in Books", f"{stats.get('missing_in_books_count', 0):,}")
    mc6.metric("Match Rate",       f"{stats.get('match_rate', 0):.1f}%")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Status legend ────────────────────────────────────────────────────────
    st.markdown("""
    <div style="display:flex; gap:12px; margin-bottom:12px; flex-wrap:wrap;">
        <span style="background:#DCFCE7; color:#166534; border-radius:20px;
               padding:3px 12px; font-size:0.75rem; font-weight:600;">● Fully Matched</span>
        <span style="background:#FEF9C3; color:#92400E; border-radius:20px;
               padding:3px 12px; font-size:0.75rem; font-weight:600;">● Near Match</span>
        <span style="background:#FEE2E2; color:#991B1B; border-radius:20px;
               padding:3px 12px; font-size:0.75rem; font-weight:600;">● Missing in Books</span>
        <span style="background:#FFF7ED; color:#9A3412; border-radius:20px;
               padding:3px 12px; font-size:0.75rem; font-weight:600;">● Missing in GSTR-2B</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Build display table ──────────────────────────────────────────────────
    SHOW_COLS = ["vendor_name","gstin","invoice_number","invoice_date",
                 "taxable_value","cgst","sgst","igst","total_gst","invoice_value"]
    RENAME = {
        "vendor_name":   "Party Name",
        "gstin":         "GSTIN",
        "invoice_number":"Invoice No",
        "invoice_date":  "Invoice Date",
        "taxable_value": "Taxable Value",
        "cgst":          "CGST",
        "sgst":          "SGST",
        "igst":          "IGST",
        "total_gst":     "Total GST",
        "invoice_value": "Invoice Value",
    }

    # Map Status to emoji label (works in both light & dark themes)
    STATUS_EMOJI = {
        "Fully Matched":     "\u2705 Fully Matched",
        "Near Match":        "\U0001f7e1 Near Match",
        "Missing in Books":  "\U0001f534 Missing in Books",
        "Missing in GSTR-2B":"\U0001f7e0 Missing in GSTR-2B",
    }

    disp_cols = ["Status"] + [c for c in SHOW_COLS if c in master_df.columns]
    if "Match Reason" in master_df.columns:
        disp_cols.append("Match Reason")
    if "Similarity %" in master_df.columns:
        disp_cols.append("Similarity %")

    disp = master_df[disp_cols].copy()
    # Apply emoji prefix to Status
    disp["Status"] = disp["Status"].map(STATUS_EMOJI).fillna(disp["Status"])
    disp = disp.rename(columns=RENAME)

    # Format dates
    if "Invoice Date" in disp.columns:
        disp["Invoice Date"] = disp["Invoice Date"].astype(str).str[:10].str.replace(" 00:00:00","",regex=False)

    # Native st.dataframe — no custom Styler (works in light & dark modes)
    st.dataframe(disp, use_container_width=True, hide_index=True, height=460)
    st.caption(
        f"**{len(master_df):,}** total records \xa0|\xa0 "
        f"GSTR-2B: **{st.session_state.get('gstr_name', '?')}** \xa0\xb7\xa0 "
        f"Purchase Register: **{st.session_state.get('pr_name', '?')}**"
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Download Excel Report ────────────────────────────────────────────────
    _, dl_col, _ = st.columns([2, 6, 2])
    with dl_col:
        try:
            from modules.reports import _build_sheet_excel
            excel_bytes = _build_sheet_excel(recon_results)
            fname = f"GST_Recon_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                "⬇  Download Reconciled Report (Excel)",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_report",
            )
        except Exception as exc:
            st.error(f"Report generation failed: {exc}")
            logger.exception("Report generation error")

# ===========================================================================
# Footer
# ===========================================================================
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown("""
<div style="text-align:center; color:#94A3B8; font-size:0.74rem;
     border-top:1px solid #E2E8F0; padding-top:16px;">
    GST Reconciliation App v2.0 &nbsp;&middot;&nbsp; AI-Powered Invoice Matching<br>
    Prepared &amp; Developed by
    <strong style="color:#4F46E5;">Karthik LVN</strong>
    &nbsp;&middot;&nbsp; &#128222; 9849270702
</div>
""", unsafe_allow_html=True)
