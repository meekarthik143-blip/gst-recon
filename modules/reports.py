"""
GST Reconciliation Report Builder
Prepared & Developed by Karthik LVN
"""

import io
import datetime

import pandas as pd
import streamlit as st

from modules.utils import setup_logging
from modules.audit import log_event

logger = setup_logging()



def _build_sheet_excel(recon_results: dict) -> bytes:
    """
    Multi-sheet Excel report.
    Sheets: Summary | Matched | Missing in Books | Missing in GSTR-2B |
            Near Match - Review (full PR+2B side-by-side + reason cols) |
            Duplicates | Unreconciled | All Records |
            As Per Books | As Per GSTR-2B
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime as dt

    # ── Styles ─────────────────────────────────────────────────────────────
    WHITE_BG   = PatternFill("solid", fgColor="FFFFFF")
    ALT_BG     = PatternFill("solid", fgColor="F8FAFC")
    BORDER_CLR = "D1D5DB"
    DATA_FONT  = Font(name="Calibri", size=9, color="111827")
    CTR  = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left",   vertical="center")

    HDR_CONFIGS = {
        "matched": ("1F5E40", "FFFFFF"), "books":   ("7F1D1D", "FFFFFF"),
        "gstr":    ("78350F", "FFFFFF"), "near":    ("78350F", "FBBF24"),
        "dup":     ("3B0764", "FFFFFF"), "unrecon": ("1E293B", "FFFFFF"),
        "all":     ("0F172A", "00D4FF"), "pr_view": ("1E3A8A", "FFFFFF"),
        "gb_view": ("14532D", "FFFFFF"),
    }

    def thin_border():
        s = Side(style="thin", color=BORDER_CLR)
        return Border(left=s, right=s, top=s, bottom=s)

    def _fmt_val(val):
        try:
            if pd.isna(val):
                return ""
        except Exception:
            pass
        if hasattr(val, "strftime"):
            return val.strftime("%d-%m-%Y")
        if isinstance(val, str) and val.endswith(" 00:00:00"):
            return val[:10]
        if isinstance(val, float):
            return round(val, 2)
        return val

    def _col_label(c):
        c = c.replace("_pr", " (Books)").replace("_gstr2b", " (2B)")
        return c.replace("_", " ").strip().title()

    def _write_rows(ws, df_out, hdr_fill, hdr_font):
        col_labels = [_col_label(c) for c in df_out.columns]
        for ci, label in enumerate(col_labels, 1):
            cell = ws.cell(1, ci, label)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = CTR; cell.border = thin_border()
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"
        for ri, row in enumerate(df_out.itertuples(index=False), 2):
            fill = WHITE_BG if ri % 2 == 0 else ALT_BG
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci)
                c.value = _fmt_val(val); c.font = DATA_FONT
                c.fill = fill; c.border = thin_border()
                c.alignment = CTR if isinstance(val, (int, float)) else LEFT
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    c.number_format = "#,##0.00"
        # Auto-fit
        for ci, label in enumerate(col_labels, 1):
            vals = [str(ws.cell(r, ci).value or "") for r in range(1, min(len(df_out) + 2, 52))]
            ws.column_dimensions[get_column_letter(ci)].width = min(max((max(len(v) for v in vals) if vals else 10) + 2, 12), 42)

    # Standard sheet (keeps only PR or plain invoice cols)
    def write_sheet(wb, title, df, tab_color, fill_key):
        ws = wb.create_sheet(title)
        ws.sheet_properties.tabColor = tab_color
        if df is None or (hasattr(df, "empty") and df.empty):
            ws["A1"] = f"No {title} records."
            ws["A1"].font = Font(name="Calibri", color="94A3B8", italic=True, size=10)
            return
        hdr_bg, hdr_fg = HDR_CONFIGS.get(fill_key, ("0F172A", "FFFFFF"))
        WANT = ["vendor_name", "gstin", "invoice_number", "invoice_date",
                "taxable_value", "cgst", "sgst", "igst", "cess", "total_gst", "invoice_value"]
        pr_cols = [c + "_pr" for c in WANT if c + "_pr" in df.columns]
        plain   = [c for c in WANT if c in df.columns and c + "_pr" not in df.columns]
        use_cols = pr_cols if pr_cols else (plain if plain else [c for c in df.columns if not c.startswith("_")][:14])
        df_out = df[use_cols].copy()
        _write_rows(ws, df_out, PatternFill("solid", fgColor=hdr_bg), Font(name="Calibri", bold=True, color=hdr_fg, size=9))

    # Near Match sheet — FULL side-by-side PR + GSTR-2B + reason cols first
    def write_near_match_sheet(wb, df, tab_color):
        ws = wb.create_sheet("Near Match - Review")
        ws.sheet_properties.tabColor = tab_color
        if df is None or (hasattr(df, "empty") and df.empty):
            ws["A1"] = "No Near Match records."
            ws["A1"].font = Font(name="Calibri", color="94A3B8", italic=True, size=10)
            return
        hdr_bg, hdr_fg = HDR_CONFIGS["near"]
        # Reason/score cols first, then all PR cols, then all GSTR-2B cols
        REASON_COLS = ["Match Reason", "Similarity %", "fuzzy_score", "confidence", "match_tier", "match_key"]
        WANT = ["vendor_name", "gstin", "invoice_number", "invoice_date",
                "taxable_value", "cgst", "sgst", "igst", "cess", "total_gst", "invoice_value"]
        pr_cols  = [c + "_pr"     for c in WANT if c + "_pr"     in df.columns]
        gst_cols = [c + "_gstr2b" for c in WANT if c + "_gstr2b" in df.columns]
        reason_cols = [c for c in REASON_COLS if c in df.columns]
        # Also include source cols if present
        extra_source = [c for c in ["source_pr"] if c in df.columns]
        full_cols = reason_cols + pr_cols + extra_source + gst_cols
        if not full_cols:
            full_cols = [c for c in df.columns if not c.startswith("_")]
        df_out = df[full_cols].copy()
        _write_rows(ws, df_out, PatternFill("solid", fgColor=hdr_bg), Font(name="Calibri", bold=True, color=hdr_fg, size=9))

    # "As Per Books" — all PR-side records with status
    def write_per_books(wb, matched, books, gstr, fuzzy):
        ws = wb.create_sheet("As Per Books")
        ws.sheet_properties.tabColor = "1E40AF"
        hdr_bg, hdr_fg = HDR_CONFIGS["pr_view"]
        WANT = ["vendor_name", "gstin", "invoice_number", "invoice_date",
                "taxable_value", "cgst", "sgst", "igst", "cess", "total_gst", "invoice_value"]

        def _extract(df, status_label, suffix="_pr"):
            if df is None or (hasattr(df, "empty") and df.empty):
                return pd.DataFrame()
            cols = [c + suffix for c in WANT if c + suffix in df.columns]
            if not cols:
                cols = [c for c in WANT if c in df.columns]
                suffix = ""
            if not cols:
                return pd.DataFrame()
            out = df[cols].copy()
            out.columns = [c.replace(suffix, "") for c in out.columns]
            out.insert(0, "Status", status_label)
            return out

        # Books records = PR side of matched + PR only unmatched (missing_in_gstr2b) + PR side of fuzzy
        frames = []
        frames.append(_extract(matched, "Matched"))
        # missing_in_gstr2b is already PR-perspective (no suffix)
        if gstr is not None and not gstr.empty:
            cols = [c for c in WANT if c in gstr.columns]
            if cols:
                out = gstr[cols].copy(); out.insert(0, "Status", "Not in GSTR-2B"); frames.append(out)
        frames.append(_extract(fuzzy, "Near Match (Verify)", suffix="_pr"))
        all_pr = pd.concat([f for f in frames if not f.empty], ignore_index=True)
        if all_pr.empty:
            ws["A1"] = "No Books records available."
            return
        _write_rows(ws, all_pr, PatternFill("solid", fgColor=hdr_bg), Font(name="Calibri", bold=True, color=hdr_fg, size=9))

    # "As Per GSTR-2B" — all GSTR-2B-side records with status
    def write_per_gstr2b(wb, matched, books, gstr, fuzzy):
        ws = wb.create_sheet("As Per GSTR-2B")
        ws.sheet_properties.tabColor = "166534"
        hdr_bg, hdr_fg = HDR_CONFIGS["gb_view"]
        WANT = ["vendor_name", "gstin", "invoice_number", "invoice_date",
                "taxable_value", "cgst", "sgst", "igst", "cess", "total_gst", "invoice_value"]

        def _extract(df, status_label, suffix="_gstr2b"):
            if df is None or (hasattr(df, "empty") and df.empty):
                return pd.DataFrame()
            cols = [c + suffix for c in WANT if c + suffix in df.columns]
            if not cols:
                cols = [c for c in WANT if c in df.columns]
                suffix = ""
            if not cols:
                return pd.DataFrame()
            out = df[cols].copy()
            out.columns = [c.replace(suffix, "") for c in out.columns]
            out.insert(0, "Status", status_label)
            return out

        frames = []
        frames.append(_extract(matched, "Matched"))
        # missing_in_books is already GSTR-2B-perspective (no suffix)
        if books is not None and not books.empty:
            cols = [c for c in WANT if c in books.columns]
            if cols:
                out = books[cols].copy(); out.insert(0, "Status", "Not in Books"); frames.append(out)
        frames.append(_extract(fuzzy, "Near Match (Verify)", suffix="_gstr2b"))
        all_gb = pd.concat([f for f in frames if not f.empty], ignore_index=True)
        if all_gb.empty:
            ws["A1"] = "No GSTR-2B records available."
            return
        _write_rows(ws, all_gb, PatternFill("solid", fgColor=hdr_bg), Font(name="Calibri", bold=True, color=hdr_fg, size=9))

    # ── Data ───────────────────────────────────────────────────────────────
    matched = recon_results.get("matched",           pd.DataFrame())
    books   = recon_results.get("missing_in_books",  pd.DataFrame())
    gstr    = recon_results.get("missing_in_gstr2b", pd.DataFrame())
    fuzzy   = recon_results.get("fuzzy_candidates",  pd.DataFrame())
    pr_dup  = recon_results.get("pr_duplicates",     pd.DataFrame())
    gst_dup = recon_results.get("gstr2b_duplicates", pd.DataFrame())
    stats   = recon_results.get("stats", {})

    dup_all = pd.concat([(pr_dup  if isinstance(pr_dup,  pd.DataFrame) else pd.DataFrame()),
                         (gst_dup if isinstance(gst_dup, pd.DataFrame) else pd.DataFrame())],
                        ignore_index=True)
    unrecon_frames = [d for d in [books, gstr, fuzzy] if isinstance(d, pd.DataFrame) and not d.empty]
    unrecon = pd.concat(unrecon_frames, ignore_index=True) if unrecon_frames else pd.DataFrame()
    all_frames = ([matched] if isinstance(matched, pd.DataFrame) and not matched.empty else []) + unrecon_frames
    all_records = pd.concat(all_frames, ignore_index=True) if all_frames else pd.DataFrame()

    # ── Workbook ───────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "00D4FF"
    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 18
    ws_sum["A1"] = "GST Input Reconciliation Report"
    ws_sum["A1"].font = Font(name="Calibri", bold=True, size=16, color="0F172A")
    ws_sum["A1"].fill = PatternFill("solid", fgColor="E0F7FA")
    ws_sum["A2"] = "Prepared & Developed by Karthik LVN"
    ws_sum["A2"].font = Font(name="Calibri", size=10, color="7C3AED", italic=True)
    ws_sum["A2"].fill = PatternFill("solid", fgColor="F3E8FF")
    ws_sum["A3"] = f"Generated: {dt.date.today().strftime('%d-%m-%Y')}"
    ws_sum["A3"].font = Font(name="Calibri", size=9, color="6B7280")
    for r, label, val, bg, fg, is_hdr in [
        (5, "Category",                           "Count",                                      "1E40AF", "FFFFFF", True),
        (6, "Matched (Fully Reconciled)",          stats.get("total_matched", 0),                "ECFDF5", "065F46", False),
        (7, "Missing in Books (In 2B, not in PR)", stats.get("missing_in_books_count", 0),       "FEF2F2", "991B1B", False),
        (8, "Missing in GSTR-2B (In PR, not 2B)", stats.get("missing_in_gstr2b_count", 0),      "FFF7ED", "92400E", False),
        (9, "Near Match (Review Required)",        stats.get("fuzzy_candidates_count", 0),       "FFFBEB", "B45309", False),
        (10,"PR Duplicates",                       stats.get("pr_duplicates_count", 0),          "F5F3FF", "4C1D95", False),
        (11,"GSTR-2B Duplicates",                  stats.get("gstr2b_duplicates_count", 0),      "F5F3FF", "4C1D95", False),
        (12,"Match Rate (%)",                      f"{stats.get('match_rate', 0):.2f}%",         "EFF6FF", "1E3A8A", False),
    ]:
        lc = ws_sum.cell(r, 1, label)
        vc = ws_sum.cell(r, 2, val)
        for cell in [lc, vc]:
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.font   = Font(name="Calibri", bold=is_hdr, size=10, color=fg)
            cell.border = thin_border()
            cell.alignment = LEFT
        vc.alignment = CTR
        ws_sum.row_dimensions[r].height = 18

    # Data sheets
    write_sheet(wb, "Matched",            matched,    "34D399", "matched")
    write_sheet(wb, "Missing in Books",   books,      "F87171", "books")
    write_sheet(wb, "Missing in GSTR-2B", gstr,       "FB923C", "gstr")
    write_near_match_sheet(wb, fuzzy,     "FBBF24")
    write_sheet(wb, "Duplicates",         dup_all,    "A78BFA", "dup")
    write_sheet(wb, "Unreconciled",       unrecon,    "94A3B8", "unrecon")
    write_sheet(wb, "All Records",        all_records,"00D4FF", "all")
    write_per_books(wb, matched, books, gstr, fuzzy)
    write_per_gstr2b(wb, matched, books, gstr, fuzzy)

    buf = io.BytesIO()
    wb.save(buf)
    log_event("EXPORT", "Sheet-wise Excel report generated")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit Reports Page
# ---------------------------------------------------------------------------

def render_reports_page() -> None:
    """Render the Reports export page."""
    import datetime

    # Header + Back to Dashboard
    hdr_col, dash_col = st.columns([8, 2])
    hdr_col.markdown("<h2 style='color:#00D4FF; margin:0;'>Reports</h2>", unsafe_allow_html=True)
    if dash_col.button("Back to Dashboard", key="rep_goto_dash", use_container_width=True):
        st.session_state["current_page"] = "Dashboard"
        st.rerun()

    recon_results = st.session_state.get("recon_results")

    if recon_results is None:
        st.markdown(
            "<div style='background:rgba(248,113,113,0.08); border:1px solid #F8717133; "
            "border-radius:10px; padding:24px; text-align:center; margin-top:16px;'>"
            "<div style='color:#F87171; font-size:1.1rem; font-weight:700;'>No reconciliation data yet</div>"
            "<div style='color:#94A3B8; margin-top:8px;'>Complete the workflow: "
            "<strong>Upload Data</strong> - <strong>Column Mapping</strong> - "
            "<strong>Reconcile</strong> - then come back here.</div></div>",
            unsafe_allow_html=True,
        )
        if st.button("Go to Upload Data", type="primary", use_container_width=True, key="rep_goto_upload"):
            st.session_state["current_page"] = "Upload Data"
            st.rerun()
        return

    stats = recon_results.get("stats", {})

    # Summary KPIs
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Matched",            f"{stats.get('total_matched', 0):,}")
    c2.metric("Missing in Books",   f"{stats.get('missing_in_books_count', 0):,}")
    c3.metric("Missing in GSTR-2B", f"{stats.get('missing_in_gstr2b_count', 0):,}")
    c4.metric("Near Match",         f"{stats.get('fuzzy_candidates_count', 0):,}")
    c5.metric("Match Rate",         f"{stats.get('match_rate', 0):.1f}%")

    st.divider()

    # Excel Download
    st.markdown(
        "<div style='font-size:1rem; font-weight:700; color:#A78BFA; margin-bottom:4px;'>"
        "Download Full Reconciliation Report (Excel)</div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "9 sheets: Summary | Matched | Missing in Books | Missing in GSTR-2B | "
        "Near Match (full side-by-side + Reason/Score) | Duplicates | Unreconciled | "
        "All Records | As Per Books | As Per GSTR-2B"
    )
    try:
        excel_bytes = _build_sheet_excel(recon_results)
        fname = f"GST_Recon_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="Download Excel Report",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_excel_direct",
            type="primary",
        )
    except Exception as e:
        st.error(f"Could not prepare Excel: {e}")

    st.divider()

    # Preview tabs
    WANT = ["vendor_name", "gstin", "invoice_number", "invoice_date",
            "taxable_value", "cgst", "sgst", "igst", "cess", "total_gst", "invoice_value"]

    def _clean_show(df, label):
        if df is None or (hasattr(df, "empty") and df.empty):
            st.info(f"No {label} records.")
            return
        # Keep PR cols (or plain) for display
        pr_cols = [c + "_pr" for c in WANT if c + "_pr" in df.columns]
        plain   = [c for c in WANT if c in df.columns and c + "_pr" not in df.columns]
        use = pr_cols if pr_cols else plain
        out = df[use].copy() if use else df.copy()
        out.columns = [c.replace("_pr", "").replace("_", " ").title() for c in out.columns]
        st.dataframe(out, use_container_width=True, hide_index=True)
        st.caption(f"{len(df):,} records")

    tab_m, tab_b, tab_g, tab_nm, tab_dup = st.tabs([
        f"Matched ({stats.get('total_matched', 0):,})",
        f"Missing in Books ({stats.get('missing_in_books_count', 0):,})",
        f"Missing in GSTR-2B ({stats.get('missing_in_gstr2b_count', 0):,})",
        f"Near Match - Review ({stats.get('fuzzy_candidates_count', 0):,})",
        f"Duplicates ({stats.get('pr_duplicates_count', 0) + stats.get('gstr2b_duplicates_count', 0):,})",
    ])

    with tab_m:
        st.caption("Invoices that exist in BOTH Purchase Register and GSTR-2B with matching details.")
        _clean_show(recon_results.get("matched"), "Matched")

    with tab_b:
        st.caption("Invoices present in GSTR-2B but NOT found in your Purchase Register (Books).")
        _clean_show(recon_results.get("missing_in_books"), "Missing in Books")

    with tab_g:
        st.caption("Invoices present in your Purchase Register (Books) but NOT found in GSTR-2B.")
        _clean_show(recon_results.get("missing_in_gstr2b"), "Missing in GSTR-2B")

    with tab_nm:
        st.markdown(
            "<div style='background:rgba(251,191,36,0.08); border:1px solid #FBBF2433; "
            "border-radius:8px; padding:10px 14px; margin-bottom:10px;'>"
            "<strong style='color:#FBBF24;'>Near Match (Review Required)</strong>"
            "<div style='color:#94A3B8; font-size:0.83rem; margin-top:4px;'>"
            "Found in both files with slight differences (e.g. invoice format INV-001 vs INV/001). "
            "<b>Match Reason</b> and <b>Similarity %</b> explain the difference. "
            "PR data shown first (Books), then GSTR-2B data. Verify and decide manually."
            "</div></div>",
            unsafe_allow_html=True,
        )
        nm_df = recon_results.get("fuzzy_candidates")
        if nm_df is None or (hasattr(nm_df, "empty") and nm_df.empty):
            st.info("No Near Match records.")
        else:
            # Full side-by-side: reason cols + PR cols + GSTR-2B cols
            REASON = ["Match Reason", "Similarity %", "fuzzy_score", "confidence", "match_tier"]
            pr_cols  = [c + "_pr"     for c in WANT if c + "_pr"     in nm_df.columns]
            gst_cols = [c + "_gstr2b" for c in WANT if c + "_gstr2b" in nm_df.columns]
            reason_cols = [c for c in REASON if c in nm_df.columns]
            show_cols = reason_cols + pr_cols + gst_cols
            if not show_cols:
                show_cols = [c for c in nm_df.columns if not c.startswith("_")]
            disp = nm_df[show_cols].copy()
            disp.columns = [
                c.replace("_pr", " (Books)").replace("_gstr2b", " (2B)")
                 .replace("_", " ").strip().title()
                for c in disp.columns
            ]
            st.dataframe(disp, use_container_width=True, hide_index=True)
            st.caption(f"{len(nm_df):,} near match records — verify each row manually")

    with tab_dup:
        st.caption("Duplicate invoice numbers detected within the same file.")
        pr_dup  = recon_results.get("pr_duplicates",     pd.DataFrame())
        gst_dup = recon_results.get("gstr2b_duplicates", pd.DataFrame())
        if isinstance(pr_dup, pd.DataFrame) and not pr_dup.empty:
            st.markdown("**Purchase Register Duplicates**")
            _clean_show(pr_dup, "PR Duplicates")
        if isinstance(gst_dup, pd.DataFrame) and not gst_dup.empty:
            st.markdown("**GSTR-2B Duplicates**")
            _clean_show(gst_dup, "GSTR-2B Duplicates")

    st.markdown(
        "<div style='text-align:center; color:#374151; font-size:0.75rem; margin-top:24px;'>"
        "Prepared &amp; Developed by <strong>Karthik LVN</strong> &nbsp;&#183;&nbsp; "
        "2026 All Rights Reserved</div>",
        unsafe_allow_html=True,
    )
